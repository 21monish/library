import csv
import datetime
import io
from decimal import Decimal

from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import Book, BookInstance, Loan, Penalty, Reservation, ScanAudit


def report_period(params):
    """Return validated inclusive date boundaries, defaulting to the last 90 days."""
    today = timezone.localdate()
    default_start = today - datetime.timedelta(days=89)
    try:
        start = datetime.date.fromisoformat(params.get('date_from', ''))
    except (TypeError, ValueError):
        start = default_start
    try:
        end = datetime.date.fromisoformat(params.get('date_to', ''))
    except (TypeError, ValueError):
        end = today
    if start > end:
        start, end = end, start
    return start, end


def _range(field, start, end):
    return {f'{field}__date__gte': start, f'{field}__date__lte': end}


def build_report(start, end):
    loans = Loan.objects.filter(**_range('issued_at', start, end))
    returns = Loan.objects.filter(status='returned', **_range('returned_at', start, end))
    penalties = Penalty.objects.filter(**_range('assessed_at', start, end))
    reservations = Reservation.objects.filter(**_range('created_at', start, end))
    scans = ScanAudit.objects.filter(**_range('created_at', start, end))
    today = timezone.localdate()

    monthly_rows = list(
        loans.annotate(month=TruncMonth('issued_at'))
        .values('month').annotate(total=Count('id')).order_by('month')
    )
    status_rows = list(loans.values('status').annotate(total=Count('id')).order_by('status'))
    genre_rows = list(
        Book.objects.filter(bookinstance__loans__in=loans)
        .values('genre__name').annotate(total=Count('bookinstance__loans', distinct=True))
        .order_by('-total', 'genre__name')[:8]
    )
    top_books = list(
        loans.values('book_title').annotate(total=Count('id')).order_by('-total', 'book_title')[:10]
    )
    top_members = list(
        loans.values('borrower__username', 'borrower__first_name', 'borrower__last_name')
        .annotate(total=Count('id')).order_by('-total', 'borrower__username')[:10]
    )
    penalty_totals = penalties.aggregate(
        assessed=Sum('amount'),
        outstanding=Sum('amount', filter=Q(status='unpaid')),
        collected=Sum('amount', filter=Q(status='paid')),
        waived=Sum('amount', filter=Q(status='waived')),
    )
    metrics = {
        'loans': loans.count(),
        'returns': returns.count(),
        'active': Loan.objects.filter(status='active').count(),
        'overdue': Loan.objects.filter(status='active', due_date__lt=today).count(),
        'reservations': reservations.count(),
        'scans': scans.count(),
        'scan_success_rate': round(scans.filter(success=True).count() * 100 / scans.count()) if scans.count() else 0,
        'penalties_assessed': penalty_totals['assessed'] or 0,
        'penalties_outstanding': penalty_totals['outstanding'] or 0,
        'penalties_collected': penalty_totals['collected'] or 0,
        'penalties_waived': penalty_totals['waived'] or 0,
        'inventory': BookInstance.objects.count(),
        'available': BookInstance.objects.filter(status='a').count(),
    }
    return {
        'start': start,
        'end': end,
        'metrics': metrics,
        'monthly': [{'label': row['month'].strftime('%b %Y'), 'total': row['total']} for row in monthly_rows],
        'statuses': [{'label': dict(Loan.STATUS_CHOICES).get(row['status'], row['status']), 'total': row['total']} for row in status_rows],
        'genres': [{'label': row['genre__name'] or 'Unclassified', 'total': row['total']} for row in genre_rows],
        'top_books': top_books,
        'top_members': top_members,
        'loans_queryset': loans.select_related('borrower', 'book_instance').order_by('-issued_at'),
        'penalties_queryset': penalties.select_related('borrower').order_by('-assessed_at'),
        'scans_queryset': scans.select_related('actor', 'target_user', 'book_instance').order_by('-created_at'),
    }


def _filename(extension, start, end):
    return f'shelfwise-report-{start:%Y%m%d}-{end:%Y%m%d}.{extension}'


def csv_response(report):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{_filename("csv", report["start"], report["end"])}"'
    writer = csv.writer(response)
    writer.writerow(['Shelfwise circulation report', report['start'], report['end']])
    writer.writerow([])
    writer.writerow(['Metric', 'Value'])
    for name, value in report['metrics'].items():
        writer.writerow([name.replace('_', ' ').title(), value])
    writer.writerow([])
    writer.writerow(['Issued at', 'Book', 'Member', 'Due date', 'Returned at', 'Status'])
    for loan in report['loans_queryset']:
        writer.writerow([loan.issued_at.isoformat(), loan.book_title, loan.borrower.username, loan.due_date, loan.returned_at.isoformat() if loan.returned_at else '', loan.get_status_display()])
    return response


def xlsx_bytes(report):
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Summary'
    summary.append(['Shelfwise report', f'{report["start"]} to {report["end"]}'])
    summary.append([])
    summary.append(['Metric', 'Value'])
    for name, value in report['metrics'].items():
        summary.append([name.replace('_', ' ').title(), float(value) if isinstance(value, Decimal) else value])

    loans_sheet = workbook.create_sheet('Loans')
    loans_sheet.append(['Issued at', 'Book', 'Member', 'Due date', 'Returned at', 'Status'])
    for loan in report['loans_queryset']:
        loans_sheet.append([timezone.localtime(loan.issued_at).replace(tzinfo=None), loan.book_title, loan.borrower.username, loan.due_date, timezone.localtime(loan.returned_at).replace(tzinfo=None) if loan.returned_at else None, loan.get_status_display()])

    penalty_sheet = workbook.create_sheet('Penalties')
    penalty_sheet.append(['Assessed at', 'Member', 'Book', 'Amount (INR)', 'Days overdue', 'Status'])
    for penalty in report['penalties_queryset']:
        penalty_sheet.append([timezone.localtime(penalty.assessed_at).replace(tzinfo=None), penalty.borrower.username, penalty.book_title, penalty.amount, penalty.days_overdue, penalty.get_status_display()])

    scan_sheet = workbook.create_sheet('Scan audit')
    scan_sheet.append(['Created at', 'Actor', 'Action', 'Target member', 'Success', 'Message'])
    for scan in report['scans_queryset']:
        scan_sheet.append([timezone.localtime(scan.created_at).replace(tzinfo=None), scan.actor.username if scan.actor else '', scan.get_action_display(), scan.target_user.username if scan.target_user else '', scan.success, scan.message])

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='176B4D')
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(max(len(str(cell.value or '')) for cell in column) + 2, 45)
            sheet.column_dimensions[column[0].column_letter].width = width
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def xlsx_response(report):
    response = HttpResponse(xlsx_bytes(report), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{_filename("xlsx", report["start"], report["end"])}"'
    return response


def pdf_bytes(report):
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=16 * mm, leftMargin=16 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph('Shelfwise Library Report', styles['Title']), Paragraph(f'{report["start"]:%d %b %Y} - {report["end"]:%d %b %Y}', styles['Normal']), Spacer(1, 6 * mm)]
    metrics = [['Metric', 'Value']] + [[name.replace('_', ' ').title(), str(value)] for name, value in report['metrics'].items()]
    table = Table(metrics, colWidths=[110 * mm, 45 * mm], repeatRows=1)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#176B4D')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('GRID', (0, 0), (-1, -1), .4, colors.HexColor('#CCD8D1')), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F8F5')]), ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('PADDING', (0, 0), (-1, -1), 6)]))
    story.extend([table, Spacer(1, 7 * mm), Paragraph('Most borrowed books', styles['Heading2'])])
    books = [['Book', 'Loans']] + [[row['book_title'], row['total']] for row in report['top_books']]
    books_table = Table(books or [['No activity', '']], colWidths=[135 * mm, 20 * mm], repeatRows=1)
    books_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F2ED')), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('GRID', (0, 0), (-1, -1), .4, colors.HexColor('#CCD8D1')), ('PADDING', (0, 0), (-1, -1), 6)]))
    story.append(books_table)
    document.build(story)
    return output.getvalue()


def pdf_response(report):
    response = HttpResponse(pdf_bytes(report), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_filename("pdf", report["start"], report["end"])}"'
    return response
